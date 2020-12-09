#!/usr/bin/env python3

import os
import re
import pickle
import json
from pathlib import Path
from collections import namedtuple
from operator import itemgetter
from gw2api import GuildWars2Client

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font

import pdb
import pprint


def main():

    gw2 = build_data()
    (users, tags) = load_config()

    # creates users[user_name].achievements_dict
    populate_achievements(users)

    create_xls(gw2, users, tags)

    print("Finished.")

def build_data(file="gw2.pickle"):

    # create client.
    client = GuildWars2Client()

    # fetch current build.
    curr_build = client.build.get()

    gw2 = {
        'build': None,
        'groups': {},
        'categories': {},
        'achievements': {}
    }

    # load the save file.. if we can.
    try:
        with open(file, 'rb') as fh:
            gw2 = pickle.load(fh)
    except FileNotFoundError:
        print('No {} file found.'.format(file))
        pass

    if gw2['build'] != curr_build:

        print('New GW2 version found.')

        # update the build
        gw2['build'] = curr_build

        # determine what needs to be updated.
        curr_grp = client.achievementsgroups.get()
        curr_cat = client.achievementscategories.get()
        curr_ach = client.achievements.get()

        # ####
        # Groups
        # ####

        # check and update groups.
        if sorted(gw2['groups'].keys()) != sorted(curr_grp):

            # create list of items missing from gw2.
            fetch_groups = curr_grp - gw2['groups'].keys()

            if fetch_groups:

                print('found {} new Achievement Groups.'.format(len(fetch_groups)))

                # pull data from Anet.
                data = client.achievementsgroups.get(ids=fetch_groups)

                # set the new category.
                for item in data:
                    print('Adding "{}"'.format(item['name']))
                    gw2['groups'][item['id']] = item

        # ####
        # Categories
        # ####

        # check and update categories.
        if sorted(gw2['categories'].keys()) != sorted(curr_cat):

            # create list of items missing from gw2.
            fetch_cats = curr_cat - gw2['categories'].keys()

            if fetch_cats:

                print('found {} new Achievement Categories.'.format(len(fetch_cats)))

                # pull data from Anet.
                data = client.achievementscategories.get(ids=fetch_cats)
                # set the new category.
                for item in data:
                    print('Adding "{}"'.format(item['name']))
                    gw2['categories'][item['id']] = item

        # ####
        # Achievements
        # ####

        if sorted(gw2['achievements'].keys()) != sorted(curr_ach):

            # create list of items missing from gw2.
            fetch_ach = curr_ach - gw2['achievements'].keys()

            if fetch_ach:

                print('found {} new Achievements.'.format(len(fetch_ach)))

                # GW2 API restricts id lists to 200 items.
                # lets break this into chunks.
                chunks = [list(fetch_ach)[x:x+200] for x in range(0, len(fetch_ach),200)]

                for fetch_chunk in chunks:

                    # pull data from Anet.
                    data = client.achievements.get(ids=fetch_chunk)

                    # set the new category.
                    for item in data:
                        print('Adding "{}"'.format(item['name']))
                        gw2['achievements'][item['id']] = item

        try:
            # Save the data to file.
            with open(file, 'wb') as fh:
                pickle.dump(gw2, fh)
                fh.close()

            with open('gw2.json', 'w') as fh:

                # dump a json version fo debug usage.
                json.dump(gw2, fh, indent=4)
                fh.close()
        except:
            pass
    # return data.
    return(gw2)

def load_config(file="config.json"):

    users = {}
    tags = {}

    conf = Path(file)
    if conf.is_file() is False:
        print("ERROR: {} is missing/unreadable.".format(file))
        exit(1)

    # open inputs.
    with open(file) as fh:
        data = json.load(fh)

        if 'Users' in data:

            """
                "Users": {
                    "User1": "API_KEY",
                    "User2": "API_KEY",
                },
            """

            for user, api_key in data['Users'].items():
                client = GuildWars2Client(verify_ssl=False, api_key=api_key)
                users[user] = client
        else :
            print("ERROR: No 'Users' entry in config.")
            exit(1)

        if 'Tags' in data:

            """
                'Tags': {
                    "TagName": [
                        "Achievement1",
                        "Achievement2",
                    ],
                },
            """

            if type(data['Tags']) is dict:
                for tag in data['Tags']:
                    for achieve in data['Tags'][tag]:
                        tags[achieve] = tag

            else:
                print("ERROR: 'Tags' is not key/value list in config.")
                exit(1)

    return (users, tags)

def populate_achievements(users):

    for name in users:
        new_dict = {}

        for item in users[name].accountachievements.get():
            if ( isinstance(item, dict) is True):
                new_dict[item['id']] = item
            else:
                item = users[name].accountachievements.get()
                print('User {} API key issue: {}.'.format(name, item["text"]))

        users[name].achievements_dict = new_dict

    return()

def create_xls(gw2, users, tags, file="achievements.csv"):

    # create workbook.
    wb = Workbook()

    # set active  worksheet.
    curr_ws = wb.active

    # insert text describing this project and how to read it.
    instructions = [
        '',
        'Guild Achievements tracker:',
        '',
        'Achievment Groups are seperated as worksheets.',
        '',
        'Achievment Categories are collapsable tables.',
        '',
        'The \'Tagged\' column allows you to filter the view to only those items.']

    for line in instructions:
        curr_ws.append([line])

    skip_groups = [ 'Daily' ]

    # create the Table header.
    header = ['Title', 'Mastery']
    for name in sorted(users):
        header.append(str(name))
    header.append('Tagged')

    table_style = TableStyleInfo(name="TableStyleMedium16", showRowStripes=True)

    cat_num = 0
    for group in sorted(gw2['groups'].values(), key=itemgetter('order')):

        # Skip groups we don't care about.
        if group['name'] in skip_groups:
            continue

        print(">> Group: {} ({})".format(group['name'],group['id']))
        # create a new worksheet.
        curr_ws = wb.create_sheet(group['name'])
        curr_ws.append(header)
        curr_ws.freeze_panes = curr_ws['A2']
        curr_ws.column_dimensions['A'].width = 60;
#@@ this row needs to be Bold.

        # create a truncated dict only containing
        # the entries we care about for this category.
        # needed for sorting by order.
        trunc_cats = {item: gw2['categories'][item] for item in group['categories']}

        for cat in sorted(trunc_cats.values(), key=itemgetter('order')):

            cat_num += 1
            if cat['name'] is '' or cat['achievements'] == []:
                continue

            # append empty row
            curr_ws.append([''])

            print(">> Category: {} ({}) [{}]".format(cat['name'],cat['id'], cat_num))

            # Set Achievement Category.
            curr_ws.append([cat['name']])
            curr_cell = curr_ws.cell(row=curr_ws.max_row, column=1)
            curr_cell.font = Font(bold=True)

            ach_list = cat['achievements']

            trunc_achs = dict()
            for k in ach_list:
                trunc_achs[k] = gw2['achievements'][k]

            top_row = curr_ws.max_row
            top_row += 1

            table = [header]
            for ach in sorted(trunc_achs.values(), key=itemgetter('name')):

                if ach['name'] is '':
                    continue

                row = [ach['name']]

                # Add Mastery entry.
                if 'rewards' in ach:
                    mastery = list(
                            filter(lambda rewards: rewards['type'] == "Mastery", ach['rewards']))
                    if mastery != []:
                        row.append(mastery[0]['region'])
                    else:
                        row.append('')
                else:
                    row.append('')

                # Add users
                for user in sorted(users):

                    if ach['id'] in users[user].achievements_dict.keys():

                        ptr = users[user].achievements_dict[ach['id']]
                        # if 'done' == True
                        if ptr['done']:
                            row.append('DONE')

                        # elseif 'current' and 'max'
                        elif ptr['current'] and ptr['max']:
                            row.append('\'{}/{}\''.format(ptr['current'], ptr['max']))

                        # else
                        else:
                            row.append('UNKN')

                    else:
                        row.append('')


                # tagged entries.
                if ach['name'] in tags.keys():
                    row.append(tags[ach['name']])

                else:
                    row.append('')

                # add row to table.
                # we do this instead of directly appending, so we can size the table
                # for following operations.
                table.append(row)

            for row in table:
                curr_ws.append(row)


            top_left_cell = curr_ws.cell(row=top_row, column=1)
            bot_right_cell = curr_ws.cell(row=curr_ws.max_row, column=curr_ws.max_column)

            pat = re.compile(r'[\s:,`\'"]+')
            table_name = group['name'] + '_' + cat['name']
            table_name = pat.sub('_', table_name)

            table = Table(
                    displayName=pat.sub('_', table_name),
                    ref="{}:{}".format(top_left_cell.coordinate, bot_right_cell.coordinate))
            print ("Table: {} - {}:{}".format(table_name, top_left_cell.coordinate, bot_right_cell.coordinate))

            table.tableStyleInfo = table_style
            curr_ws.add_table(table)
            curr_ws.row_dimensions.group(top_row, curr_ws.max_row, hidden=False)

    # write workbook to file.
    wb.save("achievements.xlsx")

main()
